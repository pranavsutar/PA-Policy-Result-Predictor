#!/usr/bin/env python3
"""
run_pipeline.py — Orchestrator. Runs the full pipeline end-to-end or selectively.

Usage:
  # Full pipeline (uses provider from .env, default: groq-8b-focused)
  python run_pipeline.py --api-key YOUR_KEY

  # Explicit provider
  python run_pipeline.py --api-key gsk_... --provider groq-8b-focused
  python run_pipeline.py --api-key gsk_... --provider groq-70b-focused

  # Dry run (no API calls, test PDF extraction only)
  python run_pipeline.py --dry-run

  # Single file + brand
  python run_pipeline.py --api-key KEY --file 330109-4880941.pdf --brand TREMFYA

  # First N rows only
  python run_pipeline.py --api-key KEY --subset 5

  # Specific stage only
  python run_pipeline.py --api-key KEY --stage extract
  python run_pipeline.py --api-key KEY --stage validate
  python run_pipeline.py --stage score          # no API key needed

  # Print prompt without calling API
  python run_pipeline.py --file 330109-4880941.pdf --brand TREMFYA --prompt-only

  # Re-run without cache
  python run_pipeline.py --api-key KEY --no-cache

  # Re-extract and diff against cached result
  python run_pipeline.py --api-key KEY --file 330109-4880941.pdf --brand TREMFYA --compare-cache
"""
import argparse
import json
import os
import sys

import openpyxl

from config import (
    get_pdf_dir, get_excel_path, SUBMISSION_COLUMNS, CACHE_DIR, OUTPUT_DIR,
    LLM_PROVIDER, GROQ_API_KEYS,
)
from pdf_extractor import process_all_pdfs, extract_pdf_text, build_context_package
from prompts import (
    SYSTEM_PROMPT, build_extraction_prompt, select_few_shot_examples,
    format_few_shot_examples,
)
from extractor import StandardExtractor, OptimizedGroqExtractor, AllKeysExhausted
from scorer import compute_access_score, check_score_sanity
from validator import (
    normalize_output, enforce_business_rules,
    validate_cross_row_consistency, validate_distributions,
    validate_row_completeness, build_final_csv,
)


def load_submission_rows(excel_path: str) -> list[tuple[str, str]]:
    """Load (filename, brand) pairs from the Submissions tab."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Submissions"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and row[1]:
            rows.append((str(row[0]).strip(), str(row[1]).strip()))
    return rows


def filter_rows(rows: list[tuple[str, str]], filename: str | None,
                brand: str | None, subset: int | None) -> list[tuple[str, str]]:
    """Filter submission rows based on CLI flags."""
    if filename:
        rows = [(f, b) for f, b in rows if f == filename]
        if brand:
            rows = [(f, b) for f, b in rows if b == brand]
        if not rows:
            print(f"ERROR: No matching rows for file={filename}, brand={brand}")
            sys.exit(1)
    if subset:
        rows = rows[:subset]
    return rows


def run_dry_run(submission_rows, filter_level=None):
    """Stage 1 & 2 only. No API calls. Print context package summaries."""
    print(f"\n{'='*70}")
    print(f"DRY RUN — Processing {len(submission_rows)} rows (no API calls)")
    print(f"{'='*70}\n")

    packages = process_all_pdfs(submission_rows, filter_level=filter_level)

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"Total packages: {len(packages)}")

    # Stats
    doc_types = {}
    pref_statuses = {}
    total_chars = 0
    for pkg in packages:
        doc_types[pkg.document_type] = doc_types.get(pkg.document_type, 0) + 1
        pref_statuses[pkg.preferred_status] = pref_statuses.get(pkg.preferred_status, 0) + 1
        total_chars += len(pkg.full_relevant_text)

    print(f"Document types: {doc_types}")
    print(f"Preferred statuses: {pref_statuses}")
    print(f"Total text to send: {total_chars:,} chars (~{total_chars//4:,} tokens)")
    print(f"Avg per row: {total_chars//len(packages):,} chars (~{total_chars//len(packages)//4:,} tokens)")

    return packages


def run_prompt_only(submission_rows):
    """Print the exact prompt that would be sent for each row."""
    pdf_dir = get_pdf_dir()
    for filename, brand in submission_rows:
        pages = extract_pdf_text(os.path.join(pdf_dir, filename))
        pkg = build_context_package(filename, brand, pages)

        examples = select_few_shot_examples(brand, pkg.document_type)
        few_shot_text = format_few_shot_examples(examples)
        extraction_prompt = build_extraction_prompt(pkg)
        full_prompt = f"{few_shot_text}\n\n---\n\n{extraction_prompt}"

        print(f"\n{'='*70}")
        print(f"PROMPT FOR: {filename} / {brand}")
        print(f"{'='*70}")
        print(f"\n--- SYSTEM PROMPT ---\n{SYSTEM_PROMPT}")
        print(f"\n--- USER PROMPT ({len(full_prompt)} chars) ---\n{full_prompt}")
        print(f"\n{'='*70}\n")

        # Save to file
        out_dir = os.path.join(CACHE_DIR, "..", "prompts")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"{filename}_{brand}_prompt.txt")
        with open(out_path, "w") as f:
            f.write(f"=== SYSTEM PROMPT ===\n{SYSTEM_PROMPT}\n\n")
            f.write(f"=== USER PROMPT ===\n{full_prompt}\n")
        print(f"Saved to: {out_path}")


def _make_extractor(api_key, use_cache, provider=None):
    """Create an extractor with the configured provider."""
    p = (provider or LLM_PROVIDER).lower()
    # Map UI names to internal provider values
    if p in ("groq-70b-focused", "groq-optimized"):
        return OptimizedGroqExtractor(api_key=api_key, use_cache=use_cache)
    return StandardExtractor(api_key=api_key, use_cache=use_cache, provider="groq",
                             pdf_dir=get_pdf_dir())


def run_extract(submission_rows, api_key, use_cache, provider=None,
                filter_level=None):
    """Run Pass 1 extraction only."""
    packages = process_all_pdfs(submission_rows, filter_level=filter_level)
    extractor = _make_extractor(api_key, use_cache, provider)
    extractions = extractor.extract_all(packages)
    return packages, extractions


def run_validate(submission_rows, api_key, use_cache, provider=None,
                 filter_level=None):
    """Run Pass 2 validation only (requires cached extractions)."""
    packages = process_all_pdfs(submission_rows, filter_level=filter_level)
    extractor = _make_extractor(api_key, use_cache, provider)

    # Load cached extractions
    extractions = []
    for pkg in packages:
        from extractor import _cache_key
        key = _cache_key(pkg.filename, pkg.brand, "extract")
        cached = extractor._load_cache(key)
        if cached is None:
            print(f"ERROR: No cached extraction for {pkg.filename}/{pkg.brand}. Run --stage extract first.")
            sys.exit(1)
        extractions.append(cached)

    validated = extractor.validate_all(extractions, packages)
    return packages, validated


def run_score(submission_rows, use_cache, filter_level=None):
    """Run scoring only (requires cached validated extractions)."""
    packages = process_all_pdfs(submission_rows, filter_level=filter_level)

    from pathlib import Path
    from extractor import _cache_key
    cache_dir = Path(CACHE_DIR)

    # Search all provider variants for cached results
    provider_variants = ["groq-8b-focused", "groq-70b-focused",
                         "groq-optimized", "groq", ""]

    results = []
    skipped = []
    for pkg in packages:
        data = None
        # Try validated cache first, then extraction cache
        for pass_name in ("validate", "extract"):
            for prov in provider_variants:
                key = _cache_key(pkg.filename, pkg.brand, pass_name, prov)
                cache_path = cache_dir / key
                if cache_path.exists():
                    try:
                        with open(cache_path) as f:
                            data = json.load(f)
                        break
                    except json.JSONDecodeError:
                        continue
            if data:
                break

        if data is None:
            print(f"WARNING: No cached data for {pkg.filename}/{pkg.brand}. Skipping.")
            skipped.append((pkg.filename, pkg.brand))
            continue

        row = normalize_output(data, pkg.filename, pkg.brand)
        row, violations = enforce_business_rules(row)
        row["Access Score"] = compute_access_score(data, pkg.brand)

        if violations:
            print(f"  {pkg.filename}/{pkg.brand}: {len(violations)} violations fixed")
            for v in violations:
                print(f"    - {v}")

        results.append(row)

    if skipped:
        print(f"\n  {len(skipped)} row(s) skipped (no cache). "
              f"Run extraction first for these rows.")

    return results


def run_compare_cache(submission_rows, api_key, provider=None):
    """Re-extract and diff against cached result."""
    pdf_dir = get_pdf_dir()
    extractor = _make_extractor(api_key, use_cache=False, provider=provider)

    for filename, brand in submission_rows:
        pages = extract_pdf_text(os.path.join(pdf_dir, filename))
        pkg = build_context_package(filename, brand, pages)

        # Load old cached result
        from extractor import _cache_key
        from pathlib import Path
        cache_dir = Path(CACHE_DIR)
        old_key = _cache_key(filename, brand, "extract")
        old_path = cache_dir / old_key
        old_result = None
        if old_path.exists():
            with open(old_path) as f:
                old_result = json.load(f)

        # Extract fresh (bypass cache)
        new_result = extractor.extract_single(pkg)

        print(f"\n{'='*70}")
        print(f"DIFF: {filename} / {brand}")
        print(f"{'='*70}")

        if old_result is None:
            print("No previous cache — showing new result only:")
            print(json.dumps(new_result, indent=2))
            continue

        # Compare
        all_keys = set(list(old_result.keys()) + list(new_result.keys()))
        diffs = []
        for key in sorted(all_keys):
            old_val = old_result.get(key)
            new_val = new_result.get(key)
            if str(old_val) != str(new_val):
                diffs.append((key, old_val, new_val))

        if not diffs:
            print("  No differences.")
        else:
            print(f"  {len(diffs)} differences:")
            for key, old_val, new_val in diffs:
                old_display = str(old_val)[:80]
                new_display = str(new_val)[:80]
                print(f"    {key}:")
                print(f"      OLD: {old_display}")
                print(f"      NEW: {new_display}")


def run_full_pipeline(submission_rows, api_key, use_cache, output_path,
                      provider=None, filter_level=None):
    """Run the complete pipeline: extract → validate → score → CSV."""
    all_rows = list(submission_rows)  # preserve for final validation

    # Stage 1 & 2: PDF extraction
    print(f"\n{'='*70}")
    print(f"STAGE 1 & 2: Reading PDFs + Filtering Relevant Sections")
    print(f"{'='*70}")
    packages = process_all_pdfs(submission_rows, filter_level=filter_level)

    # Stage 3 Pass 1: Extraction
    extractor = _make_extractor(api_key, use_cache, provider)
    try:
        extractions = extractor.extract_all(packages)
    except AllKeysExhausted as exc:
        print(f"\n{'!'*70}")
        print(f"PIPELINE STOPPED: {exc}")
        print(f"{'!'*70}")
        raise SystemExit(1)

    # Stage 3 Pass 2: Validation
    try:
        validated = extractor.validate_all(extractions, packages)
    except AllKeysExhausted as exc:
        print(f"\nWARNING: {exc}")
        print("Using unvalidated extractions for remaining rows.")
        validated = extractions

    # Stage 4: Post-processing + scoring
    print(f"\n{'='*70}")
    print(f"STAGE 4: Post-processing + Scoring")
    print(f"{'='*70}")

    results = []
    total_violations = 0
    scores = []
    score_warnings_total = 0
    for i, (pkg, data) in enumerate(zip(packages, validated)):
        row = normalize_output(data, pkg.filename, pkg.brand)
        row, violations = enforce_business_rules(row)
        score = compute_access_score(data, pkg.brand)
        row["Access Score"] = score
        scores.append(score)
        total_violations += len(violations)

        # Score sanity check
        score_warnings = check_score_sanity(data, score, pkg.brand)
        for w in score_warnings:
            ext._log_manual_review(pkg.filename, pkg.brand, f"Score sanity: {w}")
        score_warnings_total += len(score_warnings)

        viol_note = f", {len(violations)} rule fix(es)" if violations else ""
        warn_note = f", {len(score_warnings)} score warning(s)" if score_warnings else ""
        print(f"  [{i+1}/{len(packages)}] {pkg.brand:15s} "
              f"Score: {str(score):>3s}{viol_note}{warn_note}")

        results.append(row)

    # Flush review flags (replaces file with only this run's entries)
    ext._flush_manual_review()

    # Score distribution summary
    numeric_scores = [int(s) for s in scores if str(s).isdigit()]
    na_scores = sum(1 for s in scores if s == "NA")
    if numeric_scores:
        avg = sum(numeric_scores) / len(numeric_scores)
        buckets = {}
        for s in numeric_scores:
            buckets[s] = buckets.get(s, 0) + 1
        dist = ", ".join(f"{k}:{v}" for k, v in sorted(buckets.items()))
        print(f"\n  Avg score: {avg:.0f} | Distribution: {dist}")
    if na_scores:
        print(f"  Rows with NA score (insufficient data): {na_scores}")
    if total_violations:
        print(f"  Business rule fixes: {total_violations}")
    if score_warnings_total:
        print(f"  Score sanity warnings: {score_warnings_total} (see manual_review.txt)")

    # Validation checks
    print(f"\n{'='*70}")
    print(f"VALIDATION")
    print(f"{'='*70}")

    completeness_warnings = validate_row_completeness(results)
    cross_row_warnings = validate_cross_row_consistency(results)
    dist_warnings = validate_distributions(results)

    all_warnings = completeness_warnings + cross_row_warnings + dist_warnings
    if all_warnings:
        print(f"\n⚠️  {len(all_warnings)} warnings:")
        for w in all_warnings:
            print(f"  - {w}")
    else:
        print("\n✅ All validation checks passed.")

    # Write CSV
    csv_warnings = build_final_csv(results, output_path, expected_pairs=all_rows)
    if csv_warnings:
        for w in csv_warnings:
            print(f"  ⚠️  {w}")

    print(f"\n{'='*70}")
    print(f"DONE")
    print(f"{'='*70}")
    xlsx_path = output_path.rsplit(".", 1)[0] + ".xlsx"
    print(f"Output (CSV):  {output_path}")
    print(f"Output (XLSX): {xlsx_path}")
    print(f"Rows: {len(results)}")
    print(f"Columns: {len(SUBMISSION_COLUMNS)}")
    print(f"API calls used: {extractor.api_calls_made}")

    # Check for manual review items
    manual_review_path = os.path.join(os.path.dirname(CACHE_DIR), "manual_review.txt")
    if os.path.exists(manual_review_path):
        with open(manual_review_path) as f:
            lines = f.readlines()
        if lines:
            print(f"\n⚠️  {len(lines)} rows used fallback defaults — see {manual_review_path}")

    return results


def main():
    parser = argparse.ArgumentParser(
        description="Payer Policy Intelligence Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--api-key", default=None,
                        help="Groq API key")
    parser.add_argument("--provider",
                        choices=["groq-8b-focused", "groq-70b-focused"],
                        default=None,
                        help=f"LLM provider (default: from .env or '{LLM_PROVIDER}')")
    parser.add_argument("--file", default=None,
                        help="Process only this PDF filename")
    parser.add_argument("--brand", default=None,
                        help="Process only this brand (use with --file)")
    parser.add_argument("--subset", type=int, default=None,
                        help="Process only the first N rows")
    parser.add_argument("--stage", choices=["extract", "validate", "score"],
                        default=None, help="Run only this stage")
    parser.add_argument("--dry-run", action="store_true",
                        help="Run Stage 1 & 2 only, no API calls")
    parser.add_argument("--prompt-only", action="store_true",
                        help="Print prompts without calling API")
    parser.add_argument("--no-cache", action="store_true",
                        help="Ignore cached results, re-extract everything")
    parser.add_argument("--compare-cache", action="store_true",
                        help="Re-extract and diff against cached result")
    parser.add_argument("--output", default=None,
                        help="Output CSV path (default: output/results.csv)")
    parser.add_argument("--filter-level", choices=["page", "paragraph", "sentence"],
                        default=None,
                        help="Context filter granularity (default: from .env or 'sentence')")

    args = parser.parse_args()

    # Resolve provider
    provider = args.provider or LLM_PROVIDER

    # Resolve API key: CLI flag > .env
    api_key = args.api_key
    if not api_key and GROQ_API_KEYS:
        api_key = ",".join(GROQ_API_KEYS)

    # Resolve output path
    output_path = args.output or str(OUTPUT_DIR / "results.csv")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)

    # Load submission rows
    excel_path = get_excel_path()
    print(f"Loading submissions from: {excel_path}")
    all_submission_rows = load_submission_rows(excel_path)
    print(f"Total submission rows: {len(all_submission_rows)}")

    # Filter
    submission_rows = filter_rows(all_submission_rows, args.file, args.brand, args.subset)
    print(f"Processing: {len(submission_rows)} rows")

    use_cache = not args.no_cache
    filter_level = args.filter_level  # None means use config default

    # Route to the right mode
    if args.dry_run:
        run_dry_run(submission_rows, filter_level=filter_level)
        return

    if args.prompt_only:
        run_prompt_only(submission_rows)
        return

    if args.compare_cache:
        if not api_key:
            print(f"ERROR: --compare-cache requires an API key (--api-key or .env)")
            sys.exit(1)
        run_compare_cache(submission_rows, api_key, provider)
        return

    if args.stage == "extract":
        if not api_key:
            print(f"ERROR: --stage extract requires an API key (--api-key or .env)")
            sys.exit(1)
        run_extract(submission_rows, api_key, use_cache, provider,
                    filter_level=filter_level)
        return

    if args.stage == "validate":
        if not api_key:
            print(f"ERROR: --stage validate requires an API key (--api-key or .env)")
            sys.exit(1)
        run_validate(submission_rows, api_key, use_cache, provider,
                     filter_level=filter_level)
        return

    if args.stage == "score":
        results = run_score(submission_rows, use_cache, filter_level=filter_level)
        csv_warnings = build_final_csv(results, output_path)
        print(f"Scored {len(results)} rows → {output_path}")
        return

    # Full pipeline
    if not api_key:
        print("ERROR: Full pipeline requires an API key.")
        print("Set GROQ_API_KEYS in pipeline/.env,")
        print("or pass --api-key")
        print("Use --dry-run or --prompt-only for no-API modes.")
        sys.exit(1)

    run_full_pipeline(submission_rows, api_key, use_cache, output_path,
                      provider, filter_level=filter_level)


if __name__ == "__main__":
    main()
