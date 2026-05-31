"""
app.py — Streamlit dashboard for Payer Policy Intelligence pipeline.
Run: streamlit run app.py --server.port 8501
"""
import io, json, os, re, sys
from pathlib import Path
import streamlit as st
import pandas as pd
import warnings

warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(__file__))

from config import (
    get_pdf_dir, get_excel_path, SUBMISSION_COLUMNS,
    CACHE_DIR, CALL_SPACING,
    LLM_PROVIDER, GROQ_API_KEYS, GROQ_EXTRACTION_MODEL, GROQ_VALIDATION_MODEL,
)
from pdf_extractor import extract_pdf_text, build_context_package, process_all_pdfs
from prompts import (
    SYSTEM_PROMPT, build_extraction_prompt,
    select_few_shot_examples, format_few_shot_examples,
)
from scorer import (
    compute_access_score, compute_confidence, check_score_sanity,
    score_steps_brands, score_steps_generic,
    score_phototherapy, score_tb_test, score_age, score_init_auth_duration,
    score_reauth_required, score_quantity_limits, score_specialist,
    score_reauth_requirements,
)
from validator import (
    normalize_output, enforce_business_rules,
    validate_cross_row_consistency, validate_distributions,
    validate_row_completeness, KEY_TO_COLUMN,
)
from extractor import StandardExtractor, OptimizedGroqExtractor, AllKeysExhausted, _cache_key

# ── Theme CSS ────────────────────────────────────────────────────────────────
THEME_CSS = """
<style>
/* Neutral styles that work with Streamlit's built-in Light/Dark themes */
.status-ok { color: #059669; } .status-warn { color: #D97706; } .status-err { color: #DC2626; }
div[data-testid="stMetric"] label { font-size: 13px; }
</style>"""

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Payer Policy Intelligence", page_icon="Rx",
                   layout="wide", initial_sidebar_state="expanded")

# ── Session state defaults ───────────────────────────────────────────────────
if "extra_keys" not in st.session_state:
    st.session_state.extra_keys = []

st.markdown(THEME_CSS, unsafe_allow_html=True)

# ── Helpers ──────────────────────────────────────────────────────────────────
@st.cache_data
def load_submission_rows():
    import openpyxl
    wb = openpyxl.load_workbook(get_excel_path())
    ws = wb["Submissions"]
    return [
        (str(r[0]).strip(), str(r[1]).strip())
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
        if r[0] and r[1]
    ]


@st.fragment
def _render_download_buttons():
    """Download buttons isolated in a fragment to avoid full-app rerun."""
    dl_df = st.session_state.get("_dl_df")
    if dl_df is None or len(dl_df) == 0:
        return
    c1, c2 = st.columns(2)
    with c1:
        csv_data = dl_df.to_csv(index=False)
        st.download_button(
            "⬇️ Download CSV", csv_data,
            file_name="ppi_results.csv", mime="text/csv",
        )
    with c2:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
            dl_df.to_excel(writer, index=False, sheet_name="Results")
        st.download_button(
            "⬇️ Download Excel", buf.getvalue(),
            file_name="ppi_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def _get_active_keys():
    """Return active API keys based on the current provider."""
    keys = list(GROQ_API_KEYS)
    keys.extend(st.session_state.extra_keys)
    return [k for k in keys if k and "your-key" not in k and "your-groq" not in k]


def load_cached_results():
    """Load cached results from ALL providers.

    Priority order: current provider > other providers > legacy (no provider).
    Validated results preferred over extract-only. Adds a 'Provider' column
    so the user can see which provider produced each result.
    """
    cache_dir = Path(CACHE_DIR)
    if not cache_dir.exists():
        return []
    current = st.session_state.get("llm_provider", LLM_PROVIDER)
    all_providers = ["groq-8b-focused", "groq-70b-focused",
                     "groq-optimized", "groq"]  # include legacy names
    # Current provider first, then others, then legacy
    search_order = [current] + [p for p in all_providers if p != current] + [""]

    results = []
    for fn, br in load_submission_rows():
        data = None
        source_provider = ""
        for pn in ("validate", "extract"):
            for prov in search_order:
                p = cache_dir / _cache_key(fn, br, pn, prov)
                if p.exists():
                    try:
                        data = json.loads(p.read_text())
                        source_provider = prov or "legacy"
                        break
                    except json.JSONDecodeError:
                        continue
            if data:
                break
        if data is None:
            continue
        row = normalize_output(data, fn, br)
        row, _ = enforce_business_rules(row)
        row["Access Score"] = compute_access_score(data, br)
        row["Provider"] = source_provider
        results.append(row)
    return results


def load_raw(fn, br):
    """Load raw cached extraction for a (filename, brand) pair.

    Searches all providers: current first, then others, then legacy.
    """
    cache_dir = Path(CACHE_DIR)
    current = st.session_state.get("llm_provider", LLM_PROVIDER)
    all_providers = ["groq-8b-focused", "groq-70b-focused",
                     "groq-optimized", "groq"]
    search_order = [current] + [p for p in all_providers if p != current] + [""]
    for pn in ("validate", "extract"):
        for prov in search_order:
            p = cache_dir / _cache_key(fn, br, pn, prov)
            if p.exists():
                try:
                    return json.loads(p.read_text())
                except Exception:
                    continue
    return None


def score_breakdown_df(data, brand):
    rows = [
        ("Steps Brands", data.get("steps_through_brands"),
         score_steps_brands(data.get("steps_through_brands")), 20),
        ("Steps Generic", data.get("steps_through_generic"),
         score_steps_generic(data.get("steps_through_generic")), 15),
        ("Phototherapy", data.get("step_through_phototherapy"),
         score_phototherapy(data.get("step_through_phototherapy")), 5),
        ("TB Test", data.get("tb_test_required"),
         score_tb_test(data.get("tb_test_required")), 5),
        ("Age", data.get("age"), score_age(data.get("age"), brand), 10),
        ("Init Auth", data.get("initial_auth_duration"),
         score_init_auth_duration(data.get("initial_auth_duration")), 15),
        ("Reauth Req", f"{data.get('reauth_required')}/{data.get('reauth_duration')}",
         score_reauth_required(data.get("reauth_required"),
                               data.get("reauth_duration")), 10),
        ("Qty Limits", data.get("quantity_limits"),
         score_quantity_limits(data.get("quantity_limits")), 5),
        ("Specialist", data.get("specialist_types"),
         score_specialist(data.get("specialist_types")), 5),
        ("Reauth Reqs", str(data.get("reauth_requirements", ""))[:50],
         score_reauth_requirements(data.get("reauth_requirements")), 10),
    ]
    return pd.DataFrame(rows, columns=["Component", "Value", "Points", "Max"])


def safe_st_dataframe(df, *args, **kwargs):
    df_safe = df.copy()
    for col in df_safe.columns:
        if df_safe[col].dtype == object:
            df_safe[col] = df_safe[col].astype(str)
    return st.dataframe(df_safe, *args, **kwargs)


# ── Load data ────────────────────────────────────────────────────────────────
submission_rows = load_submission_rows()
results = load_cached_results()
pdf_dir = get_pdf_dir()
pdf_count = len([f for f in os.listdir(pdf_dir) if f.endswith(".pdf")]) if os.path.isdir(pdf_dir) else 0
all_files = sorted(set(f for f, _ in submission_rows))
all_brands = sorted(set(b for _, b in submission_rows))
active_keys = _get_active_keys()

# ── Tagged labels for PDF dropdowns ──────────────────────────────────────────
# Build {filename: [brands]} and {(filename, brand): True} for cache status
_file_brands: dict[str, list[str]] = {}
for _f, _b in submission_rows:
    _file_brands.setdefault(_f, []).append(_b)

_cached_pairs: set[tuple[str, str]] = set()
for r in results:
    _cached_pairs.add((r.get("Filename", ""), r.get("Brand", "")))


def _pdf_tag(filename: str) -> str:
    """Build a tagged label like: 330109.pdf [TREMFYA, STELARA · ✅]"""
    brands = _file_brands.get(filename, [])
    brand_str = ", ".join(brands)
    all_done = all((filename, b) in _cached_pairs for b in brands)
    any_done = any((filename, b) in _cached_pairs for b in brands)
    if all_done and brands:
        icon = "✅"
    elif any_done:
        icon = "⏳"
    else:
        icon = "◯"
    return f"{filename}  [{brand_str} · {icon}]"


def _pdf_from_tag(tag: str) -> str:
    """Extract plain filename from a tagged label."""
    return tag.split("  [")[0]


all_file_tags = [_pdf_tag(f) for f in all_files]

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Payer Policy Intelligence")
    st.caption("PA Policy Parameter Extraction")
    st.divider()

    # Provider selector
    _provider_options = ["groq-8b-focused", "groq-70b-focused"]
    _provider_index = (_provider_options.index(LLM_PROVIDER)
                       if LLM_PROVIDER in _provider_options else 0)
    provider = st.radio(
        "LLM Provider",
        _provider_options,
        index=_provider_index,
        key="llm_provider",
        horizontal=True,
    )

    if active_keys:
        st.success(f"{len(active_keys)} API key(s) active")
    else:
        st.warning("No API keys configured")

    with st.expander("Manage API Keys", expanded=not bool(active_keys)):
        new_key = st.text_input("Add key", type="password",
                                placeholder="gsk_...")
        if st.button("Add Key") and new_key.strip():
            st.session_state.extra_keys.append(new_key.strip())
            st.rerun()
        if active_keys:
            st.caption("Keys: " + ", ".join(f"...{k[-6:]}" for k in active_keys))

    if provider == "groq-70b-focused":
        st.caption(f"Model: `{GROQ_EXTRACTION_MODEL}`")
        st.caption("Split-prompt: 2 focused calls/row")
    elif provider == "groq-8b-focused":
        st.caption(f"Primary: `{GROQ_VALIDATION_MODEL}`")
        st.caption(f"Fallback: `{GROQ_EXTRACTION_MODEL}`")

    st.divider()
    st.markdown("**Context Filter**")
    filter_level = st.radio(
        "Filter level",
        ["sentence", "paragraph", "page"],
        index=0,
        key="filter_level",
        help="Controls how aggressively PDF text is filtered before sending to the LLM. "
             "'sentence' is most aggressive (recommended), 'page' sends whole pages.",
        horizontal=True,
    )

    st.divider()
    st.metric("PDFs Found", pdf_count)
    st.metric("Rows Extracted", f"{len(results)} / {len(submission_rows)}")
    cache_dir = Path(CACHE_DIR)
    cache_files = list(cache_dir.glob("*.json")) if cache_dir.exists() else []
    _current_prov = st.session_state.get("llm_provider", LLM_PROVIDER)
    _current_count = sum(1 for f in cache_files if f"_{_current_prov}." in f.name)
    _other_count = len(cache_files) - _current_count
    label = f"{_current_count}"
    if _other_count:
        label += f" (+{_other_count} other)"
    st.metric("Cached Responses", label)

    # Session usage — from the most recent pipeline run in this session
    st.divider()
    st.markdown("**Session Usage**")

    run_stats = st.session_state.get("last_run_stats")
    if run_stats:
        st.caption("Most recent run in this session:")
        tc1, tc2 = st.columns(2)
        tc1.metric("API Calls", run_stats["calls"])
        tc2.metric("Cache Hits", run_stats["cache_hits"])
        tc3, tc4 = st.columns(2)
        tc3.metric("Input", f"~{run_stats['input_tokens']:,} tok")
        tc4.metric("Output", f"~{run_stats['output_tokens']:,} tok")
        if run_stats.get("provider"):
            st.caption(f"Provider: {run_stats['provider']}")
    else:
        if cache_files:
            st.caption(f"{len(cache_files)} cached responses from previous sessions.")
        else:
            st.caption("No runs yet in this session.")

    # Estimated daily budget (not actual account usage)
    n_keys = len(active_keys)
    if n_keys:
        daily_budget = n_keys * 1500
        session_calls = run_stats["calls"] if run_stats else 0
        st.caption(f"Estimated daily budget: ~{daily_budget:,} calls "
                   f"({n_keys} key{'s' if n_keys > 1 else ''} × 1,500)")
        if session_calls:
            st.caption(f"This session: {session_calls} calls used")

    st.divider()

    # Count stale cache entries (wrong provider or legacy no-provider)
    current_provider = st.session_state.get("llm_provider", LLM_PROVIDER)
    stale_count = 0
    if cache_dir.exists():
        for f in cache_files:
            name = f.stem
            # Current provider entries contain the provider name
            if f"__{current_provider}__" not in name:
                stale_count += 1

    cc1, cc2 = st.columns(2)
    with cc1:
        if st.button("Clear Stale", help=f"Remove {stale_count} entries from other providers"):
            if cache_dir.exists():
                removed = 0
                for f in cache_dir.glob("*.json"):
                    if f"__{current_provider}__" not in f.stem:
                        f.unlink()
                        removed += 1
                st.success(f"Removed {removed} stale entries")
                st.rerun()
    with cc2:
        if st.button("Clear All"):
            if cache_dir.exists():
                for f in cache_dir.glob("*.json"):
                    f.unlink()
                st.success("Cache cleared")
                st.rerun()

# ── Tabs ─────────────────────────────────────────────────────────────────────
tab_overview, tab_results, tab_explorer, tab_run, tab_guide = st.tabs(
    ["Overview", "Results", "PDF Explorer", "Run", "Guide"]
)

# ── Overview Tab ─────────────────────────────────────────────────────────────
with tab_overview:
    if not results:
        st.markdown("### Welcome to the Pipeline Dashboard")
        st.markdown(
            "Extract Prior Authorization parameters from payer policy PDFs "
            "for Psoriasis indications and compute Access Quality scores."
        )
        st.divider()
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("PDFs", pdf_count)
        c2.metric("Rows", len(submission_rows))
        c3.metric("Keys", len(active_keys),
                  delta="Ready" if active_keys else "Missing",
                  delta_color="normal" if active_keys else "inverse")
        pct = (len(results) / len(submission_rows) * 100) if submission_rows else 0
        c4.metric("Progress", f"{pct:.0f}%")
        st.divider()
        st.markdown("### Quick Start")
        steps = [
            ("1.", "Add API Key", "Paste your API key in the sidebar", bool(active_keys)),
            ("2.", "Run Pipeline", "Go to the **Run** tab", False),
            ("3.", "View Results", "Switch to the **Results** tab", len(results) > 0),
        ]
        for icon, title, desc, done in steps:
            (st.success if done else st.info)(f"{icon} **{title}** — {desc}")
    else:
        st.markdown("### Pipeline Status")
        c1, c2, c3, c4 = st.columns(4)
        pct = len(results) / len(submission_rows) * 100
        c1.metric("Completion", f"{pct:.0f}%")
        df_temp = pd.DataFrame(results)
        # Convert Access Score to numeric once — "NA" becomes NaN
        if "Access Score" in df_temp.columns:
            df_temp["Access Score"] = pd.to_numeric(
                df_temp["Access Score"], errors="coerce")
            _valid = df_temp["Access Score"].dropna()
            if len(_valid) > 0:
                c2.metric("Avg Score", f"{_valid.mean():.0f}")
            else:
                c2.metric("Avg Score", "NA")
        c3.metric("Brands", df_temp["Brand"].nunique())
        c4.metric("Rows", f"{len(results)}/{len(submission_rows)}")
        st.divider()
        if len(results) > 0:
            st.markdown("### Brand Summary")
            brand_summary = df_temp.groupby("Brand").agg(
                Count=("Filename", "count"),
                Avg_Score=("Access Score", "mean"),
                Min_Score=("Access Score", "min"),
                Max_Score=("Access Score", "max"),
            ).round(0)
            brand_summary = brand_summary.fillna(-1).astype(int).replace(-1, "NA")
            safe_st_dataframe(brand_summary, width='stretch')

            # ── Charts ────────────────────────────────────────────────
            import altair as alt

            scores_numeric = df_temp["Access Score"].dropna()

            total_expected = len(submission_rows)
            total_processed = len(results)
            is_partial = total_processed < total_expected
            min_rows_for_charts = 5

            if len(scores_numeric) >= min_rows_for_charts:
                st.divider()
                if is_partial:
                    st.info(f"Showing charts for {total_processed} of "
                            f"{total_expected} rows "
                            f"({total_processed/total_expected*100:.0f}%). "
                            f"Run the full pipeline for complete results.")

                # Chart 1: Score Distribution
                st.markdown("### Access Score Distribution")
                st.caption("How many policies fall into each access tier. "
                           "Higher scores = fewer PA barriers for patients.")
                score_dist = (scores_numeric.astype(int)
                              .value_counts()
                              .reset_index())
                score_dist.columns = ["Score", "Count"]
                score_dist = score_dist.sort_values("Score")

                # Color scale: 0=red, 25=orange, 50=yellow, 75=light green, 100=green
                color_scale = alt.Scale(
                    domain=[0, 25, 50, 75, 100],
                    range=["#D32F2F", "#F57C00", "#FBC02D", "#7CB342", "#2E7D32"]
                )
                chart1 = (
                    alt.Chart(score_dist)
                    .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                    .encode(
                        x=alt.X("Score:O", title="Access Score",
                                sort=[0, 25, 50, 75, 100]),
                        y=alt.Y("Count:Q", title="Number of Policies"),
                        color=alt.Color("Score:O", scale=color_scale,
                                        legend=None),
                        tooltip=["Score", "Count"],
                    )
                    .properties(height=280)
                )
                # Add count labels on top of bars
                text1 = chart1.mark_text(dy=-10, fontSize=14,
                                         fontWeight="bold").encode(
                    text="Count:Q"
                )
                st.altair_chart(chart1 + text1, width='stretch')

                # Chart 2: Brand Comparison
                st.markdown("### Brand Access Comparison")
                st.caption("Average Access Score per brand, sorted by access level. "
                           "Brands on the right face fewer PA restrictions.")
                brand_scores = (df_temp.groupby("Brand")["Access Score"]
                                .mean().reset_index())
                brand_scores.columns = ["Brand", "Avg Score"]
                brand_scores = brand_scores.dropna(subset=["Avg Score"])
                brand_scores["Avg Score"] = brand_scores["Avg Score"].round(0).astype(int)
                brand_scores = brand_scores.sort_values("Avg Score")

                chart2 = (
                    alt.Chart(brand_scores)
                    .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                    .encode(
                        x=alt.X("Avg Score:Q", title="Average Access Score",
                                scale=alt.Scale(domain=[0, 100])),
                        y=alt.Y("Brand:N", title=None,
                                sort=alt.EncodingSortField(
                                    field="Avg Score", order="ascending")),
                        color=alt.Color("Avg Score:Q",
                                        scale=alt.Scale(
                                            domain=[0, 50, 100],
                                            range=["#D32F2F", "#FBC02D",
                                                   "#2E7D32"]),
                                        legend=None),
                        tooltip=["Brand", "Avg Score"],
                    )
                    .properties(height=max(200, len(brand_scores) * 28))
                )
                text2 = chart2.mark_text(dx=3, align="left",
                                          fontSize=12).encode(
                    text="Avg Score:Q"
                )
                st.altair_chart(chart2 + text2, width='stretch')

                # Chart 3: Step Therapy Burden
                st.markdown("### Step Therapy Burden by Brand")
                st.caption(
                    "Step therapy requires patients to try (and fail on) other "
                    "drugs before the prescribed biologic is approved. "
                    "Generic steps (light) = non-biologics like methotrexate. "
                    "Branded steps (dark) = other biologics used as gatekeeping. "
                    "Each step adds months of delay. This is the largest driver "
                    "of the Access Score (35 of 100 points)."
                )
                step_data = []
                for _, row in df_temp.iterrows():
                    brand = row.get("Brand", "?")
                    b_steps = row.get("Number of Steps through Brands",
                                     row.get("steps_through_brands", "NA"))
                    g_steps = row.get("Number of Steps through Generic",
                                     row.get("steps_through_generic", "NA"))
                    try:
                        b_val = int(b_steps) if str(b_steps).strip().isdigit() else 0
                    except (ValueError, TypeError):
                        b_val = 0
                    try:
                        g_val = int(g_steps) if str(g_steps).strip().isdigit() else 0
                    except (ValueError, TypeError):
                        g_val = 0
                    step_data.append({"Brand": brand,
                                      "Branded Steps": b_val,
                                      "Generic Steps": g_val})

                df_steps = pd.DataFrame(step_data)
                step_avg = (df_steps.groupby("Brand")
                            [["Branded Steps", "Generic Steps"]]
                            .mean().round(1).reset_index())
                step_avg["Total"] = step_avg["Branded Steps"] + step_avg["Generic Steps"]
                step_avg = step_avg.sort_values("Total", ascending=True)

                step_melted = step_avg.melt(
                    id_vars=["Brand"],
                    value_vars=["Branded Steps", "Generic Steps"],
                    var_name="Step Type", value_name="Avg Steps"
                )

                chart3 = (
                    alt.Chart(step_melted)
                    .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
                    .encode(
                        x=alt.X("Avg Steps:Q", title="Average Steps Required",
                                stack="zero"),
                        y=alt.Y("Brand:N", title=None,
                                sort=alt.EncodingSortField(
                                    field="Avg Steps", op="sum",
                                    order="ascending")),
                        color=alt.Color("Step Type:N",
                                        scale=alt.Scale(
                                            domain=["Branded Steps",
                                                    "Generic Steps"],
                                            range=["#1565C0", "#90CAF9"]),
                                        legend=alt.Legend(
                                            orient="top", title=None)),
                        tooltip=["Brand", "Step Type",
                                 alt.Tooltip("Avg Steps:Q", format=".1f")],
                    )
                    .properties(height=max(200, len(step_avg) * 28))
                )
                st.altair_chart(chart3, width='stretch')

            elif len(scores_numeric) > 0:
                st.divider()
                st.caption(f"Charts will appear after at least "
                           f"{min_rows_for_charts} rows are processed "
                           f"({len(scores_numeric)} so far).")

            # Quality checks — statistical sanity checks across all extracted rows.
            # These flag unusual patterns (e.g. a brand with wildly different scores
            # across payers, or a parameter that is NA for every single row).
            # They are informational, not errors.
            warns = validate_cross_row_consistency(results)
            warns += validate_distributions(results)
            if warns:
                with st.expander(
                    f"Quality Checks ({len(warns)} notes)",
                    expanded=False
                ):
                    st.caption(
                        "These are automated sanity checks on the extracted data. "
                        "They highlight statistical outliers and unusual patterns "
                        "worth a quick manual review — they do not indicate errors."
                    )
                    for w in warns:
                        st.caption(f"• {w}")

# ── Results Tab ──────────────────────────────────────────────────────────────
with tab_results:
    if not results:
        st.info("No results yet. Go to **Run** tab to extract parameters.")
    else:
        df = pd.DataFrame(results)
        fc1, fc2 = st.columns(2)
        with fc1:
            filt_file_tag = st.selectbox("Filter PDF", ["All"] + all_file_tags, key="rf")
        with fc2:
            filt_brand = st.selectbox("Filter Brand", ["All"] + all_brands, key="rb")
        filt_file = _pdf_from_tag(filt_file_tag) if filt_file_tag != "All" else "All"
        if filt_file != "All":
            df = df[df["Filename"] == filt_file]
        if filt_brand != "All":
            df = df[df["Brand"] == filt_brand]

        show_debug = st.toggle("Show Debug Panel", value=False, key="dbg")
        if show_debug:
            col_table, col_debug = st.columns([3, 2])
        else:
            col_table = st.container()
            col_debug = None

        with col_table:
            mc1, mc2, mc3 = st.columns(3)
            mc1.metric("Rows", len(df))
            if "Access Score" in df.columns and len(df) > 0:
                _sn = pd.to_numeric(df["Access Score"], errors="coerce").dropna()
                if len(_sn) > 0:
                    mc2.metric("Avg Score", f"{_sn.mean():.0f}")
                    mc3.metric("Range", f"{int(_sn.min())}-{int(_sn.max())}")
                else:
                    mc2.metric("Avg Score", "NA")
                    mc3.metric("Range", "NA")
            display_cols = [c for c in SUBMISSION_COLUMNS if c in df.columns]
            if "Provider" in df.columns:
                display_cols = display_cols + ["Provider"]
            safe_st_dataframe(df[display_cols], width='stretch', height=350)

            if len(df) > 0:
                row_opts = [f"{r['Filename']} / {r['Brand']}" for _, r in df.iterrows()]
                sel_label = st.selectbox("Inspect row", row_opts, key="insp")
                sel_fn, sel_br = sel_label.split(" / ", 1)
                sel_row = df[(df["Filename"] == sel_fn) & (df["Brand"] == sel_br)].iloc[0]

                with st.expander("Parameters", expanded=True):
                    pc1, pc2 = st.columns(2)
                    pcols = [c for c in SUBMISSION_COLUMNS
                             if c not in ("Filename", "Brand", "Access Score")]
                    half = len(pcols) // 2
                    with pc1:
                        for c in pcols[:half]:
                            st.text(f"{c}: {sel_row.get(c, '—')}")
                    with pc2:
                        for c in pcols[half:]:
                            st.text(f"{c}: {sel_row.get(c, '—')}")

                with st.expander("Score Breakdown"):
                    raw = load_raw(sel_fn, sel_br)
                    if raw:
                        bd = score_breakdown_df(raw, sel_br)
                        safe_st_dataframe(bd, width='stretch', hide_index=True)
                        actual = compute_access_score(raw, sel_br)
                        if actual == "NA":
                            st.metric("Access Score", "NA")
                        else:
                            st.metric("Access Score", f"{actual} / 100")
                        st.caption("Layered model — component points are indicators, "
                                   "not additive.")

        if show_debug and col_debug is not None and len(df) > 0:
            with col_debug:
                st.markdown(f"**Debug: {sel_fn} / {sel_br}**")
                pdf_path = os.path.join(get_pdf_dir(), sel_fn)
                if os.path.isfile(pdf_path):
                    from pdf_extractor import score_page_relevance
                    pages = extract_pdf_text(pdf_path)
                    pkg = build_context_package(sel_fn, sel_br, pages,
                                                filter_level=st.session_state.get("filter_level", "sentence"))
                    dtabs = st.tabs(["Pages", "Context", "Prompt", "Raw JSON"])
                    with dtabs[0]:
                        pg_data = []
                        for i, t in enumerate(pages):
                            sc, cats = score_page_relevance(t, sel_br)
                            pg_data.append({"Pg": i + 1, "Score": sc,
                                            "Tags": ", ".join(sorted(cats)) or "—",
                                            "Chars": len(t)})
                        safe_st_dataframe(pd.DataFrame(pg_data),
                                          width='stretch', height=250)
                        st.caption(f"Type: {pkg.document_type} · "
                                   f"Preferred: {pkg.preferred_status} · "
                                   f"Used: {pkg.relevant_pages_used}/{pkg.total_pages}")
                    with dtabs[1]:
                        st.text_area("Text sent to LLM",
                                     pkg.full_relevant_text, height=250)
                        if pkg.universal_criteria_text:
                            st.success(f"Universal criteria: "
                                       f"{len(pkg.universal_criteria_text)} chars")
                    with dtabs[2]:
                        examples = select_few_shot_examples(sel_br, pkg.document_type)
                        fst = format_few_shot_examples(examples)
                        ep = build_extraction_prompt(pkg)
                        full_p = f"{fst}\n\n---\n\n{ep}"
                        st.text_area("Full prompt", full_p, height=250)
                        st.caption(f"~{(len(SYSTEM_PROMPT) + len(full_p)) // 4} tokens")
                    with dtabs[3]:
                        raw = load_raw(sel_fn, sel_br)
                        if raw:
                            st.json(raw)
                        else:
                            st.info("No cached extraction")

        st.divider()
        if len(df) > 0:
            # Store for fragment access (avoids full-app rerun on download click)
            st.session_state["_dl_df"] = df[display_cols].copy()
            _render_download_buttons()

# ── PDF Explorer Tab ─────────────────────────────────────────────────────────
with tab_explorer:
    st.markdown("### PDF Explorer")
    st.caption("Deep-dive into how the pipeline processes each PDF — page scoring, "
               "context filtering, and extraction results.")

    # ── Selectors ────────────────────────────────────────────────────────────
    ex_c1, ex_c2 = st.columns(2)
    with ex_c1:
        ex_file_tag = st.selectbox("Select PDF", all_file_tags, key="ex_f")
        ex_file = _pdf_from_tag(ex_file_tag)
    with ex_c2:
        ex_brands = _file_brands.get(ex_file, ["TREMFYA"])
        ex_brand = st.selectbox("Brand", ex_brands, key="ex_b")

    pdf_path = os.path.join(get_pdf_dir(), ex_file)
    if not os.path.isfile(pdf_path):
        st.error(f"PDF not found: {ex_file}")
    else:
        from pdf_extractor import (score_page_relevance, classify_document,
                                   detect_preferred_status, extract_universal_criteria)

        pages = extract_pdf_text(pdf_path)
        pkg = build_context_package(ex_file, ex_brand, pages,
                                    filter_level=st.session_state.get("filter_level", "sentence"))

        # ── Section 1: Document Profile ──────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Document Profile")
        pc1, pc2, pc3, pc4, pc5 = st.columns(5)
        pc1.metric("Pages", len(pages))
        pc2.metric("Type", pkg.document_type)
        pc3.metric("Preferred", pkg.preferred_status)
        pc4.metric("Pages Used", f"{pkg.relevant_pages_used}/{pkg.total_pages}")
        file_kb = os.path.getsize(pdf_path) / 1024
        pc5.metric("File Size", f"{file_kb:.0f} KB")

        brands_in_file = _file_brands.get(ex_file, [])
        cached_brands = [b for b in brands_in_file if (ex_file, b) in _cached_pairs]
        pending_brands = [b for b in brands_in_file if (ex_file, b) not in _cached_pairs]

        info_parts = [f"**Brands:** {', '.join(brands_in_file)}"]
        if cached_brands:
            info_parts.append(f"**Extracted:** {', '.join(cached_brands)}")
        if pending_brands:
            info_parts.append(f"**Pending:** {', '.join(pending_brands)}")
        st.markdown(" · ".join(info_parts))

        # ── Section 2: Page Relevance ────────────────────────────────────────
        st.markdown("---")
        cur_filter = st.session_state.get("filter_level", "sentence")
        st.markdown("#### Page Relevance Scores")
        st.caption(f"Keyword relevance score per page. "
                   f"Actual content selection uses **{cur_filter}**-level filtering.  \n"
                   f"Dark green = strong match (≥8) · "
                   f"Light green = moderate (≥3) · "
                   f"Yellow = weak (>0) · "
                   f"Gray = no match")

        page_data = []
        for i, t in enumerate(pages):
            sc, cats = score_page_relevance(t, ex_brand)
            cat_str = ", ".join(sorted(cats)) if cats else "—"
            preview = t[:120].replace("\n", " ").strip()
            page_data.append({
                "Page": i + 1,
                "Score": round(sc, 1),
                "Categories": cat_str,
                "Chars": len(t),
                "Preview": preview[:80] + "..." if len(preview) > 80 else preview,
            })

        df_pages = pd.DataFrame(page_data)

        def _color_row(row):
            sc = row["Score"]
            if sc >= 8:
                return ["background-color: rgba(76,175,80,0.3)"] * len(row)
            elif sc >= 3:
                return ["background-color: rgba(76,175,80,0.15)"]  * len(row)
            elif sc > 0:
                return ["background-color: rgba(255,193,7,0.15)"] * len(row)
            return ["opacity: 0.4"] * len(row)

        styled = df_pages.style.apply(_color_row, axis=1)
        st.dataframe(styled, width='stretch',
                     height=min(400, 35 * len(df_pages) + 38))

        raw_chars = sum(len(p) for p in pages)
        filt_chars = len(pkg.full_relevant_text)
        reduction = (1 - filt_chars / raw_chars) * 100 if raw_chars > 0 else 0

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Total Pages", len(pages))
        mc2.metric("Pages Contributing", pkg.relevant_pages_used)
        mc3.metric("Raw Text", f"{raw_chars:,} chars")
        mc4.metric("After Filter", f"{filt_chars:,} chars",
                   delta=f"-{reduction:.0f}%", delta_color="normal")

        # ── Section 3: Context Sent to LLM ──────────────────────────────────
        st.markdown("---")
        st.markdown("#### Context Sent to LLM")

        ctx_tabs = st.tabs(["Full Context", "Tier Breakdown", "Prompt Preview"])

        with ctx_tabs[0]:
            token_est = len(pkg.full_relevant_text) // 4
            st.caption(f"{len(pkg.full_relevant_text):,} chars · ~{token_est:,} tokens "
                       f"(filter: {cur_filter})")
            st.text_area("Text sent to LLM (after filtering + cleanup)",
                         pkg.full_relevant_text, height=300, key="ex_ctx")

        with ctx_tabs[1]:
            # Show each tier's contribution
            if pkg.universal_criteria_text:
                with st.expander(f"Universal Criteria — "
                                 f"{len(pkg.universal_criteria_text):,} chars",
                                 expanded=False):
                    st.text_area("Universal criteria text",
                                 pkg.universal_criteria_text, height=200,
                                 key="ex_univ")
            else:
                st.caption("No universal criteria section detected "
                           "(single-drug or flat-catalog document).")

            if pkg.psoriasis_section_text:
                with st.expander(f"PsO + Brand Section — "
                                 f"{len(pkg.psoriasis_section_text):,} chars",
                                 expanded=True):
                    st.text_area("Psoriasis-specific content",
                                 pkg.psoriasis_section_text, height=200,
                                 key="ex_pso")
            else:
                st.caption("No distinct psoriasis section found.")

            if pkg.reauthorization_text:
                with st.expander(f"Reauthorization — "
                                 f"{len(pkg.reauthorization_text):,} chars",
                                 expanded=False):
                    st.text_area("Reauthorization text",
                                 pkg.reauthorization_text, height=150,
                                 key="ex_reauth")

            if pkg.quantity_limit_text:
                with st.expander(f"Quantity Limits — "
                                 f"{len(pkg.quantity_limit_text):,} chars",
                                 expanded=False):
                    st.text_area("Quantity limit text",
                                 pkg.quantity_limit_text, height=150,
                                 key="ex_ql")

        with ctx_tabs[2]:
            examples = select_few_shot_examples(ex_brand, pkg.document_type)
            fst = format_few_shot_examples(examples)
            ep = build_extraction_prompt(pkg)
            full_prompt = f"{fst}\n\n---\n\n{ep}"
            total_tokens = (len(SYSTEM_PROMPT) + len(full_prompt)) // 4
            st.caption(f"System: ~{len(SYSTEM_PROMPT)//4:,} tokens · "
                       f"User: ~{len(full_prompt)//4:,} tokens · "
                       f"Total: ~{total_tokens:,} tokens")
            st.text_area("Full prompt (copy for AI Studio)",
                         full_prompt, height=300, key="ex_prompt")

        # ── Section 4: Extraction Result ─────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Extraction Result")

        raw = load_raw(ex_file, ex_brand)
        if raw is None:
            st.info(f"No cached extraction for {ex_file} / {ex_brand}. "
                    f"Run the pipeline first.")
        else:
            # Parameters side by side
            param_c1, param_c2 = st.columns(2)
            row_data = normalize_output(raw, ex_file, ex_brand)
            row_data, violations = enforce_business_rules(row_data)
            row_data["Access Score"] = compute_access_score(raw, ex_brand)

            pcols = [c for c in SUBMISSION_COLUMNS
                     if c not in ("Filename", "Brand")]
            half = len(pcols) // 2
            with param_c1:
                for c in pcols[:half]:
                    val = row_data.get(c, "—")
                    st.markdown(f"**{c}**")
                    st.text(str(val))
            with param_c2:
                for c in pcols[half:]:
                    val = row_data.get(c, "—")
                    st.markdown(f"**{c}**")
                    st.text(str(val))

            # Score breakdown
            st.markdown("##### Score Breakdown")
            bd = score_breakdown_df(raw, ex_brand)
            total_score = compute_access_score(raw, ex_brand)
            sc1, sc2 = st.columns([3, 1])
            with sc1:
                safe_st_dataframe(bd, hide_index=True, width="stretch")
                st.caption("Component points are indicators. Final score uses "
                           "a layered model with hard floors and interaction penalties.")
            with sc2:
                if total_score == "NA":
                    st.metric("Access Score", "NA")
                    st.error("Insufficient data")
                else:
                    st.metric("Access Score", f"{total_score} / 100")
                    if total_score >= 70:
                        st.success("Good access")
                    elif total_score >= 40:
                        st.warning("Moderate access")
                    else:
                        st.error("Restricted access")

            # Reasoning + corrections
            reasoning = raw.get("reasoning", "")
            if reasoning:
                st.markdown("##### LLM Reasoning")
                st.info(reasoning)

            corrections = raw.get("corrections")
            if corrections:
                st.markdown("##### Validation Corrections (Pass 2)")
                if isinstance(corrections, list):
                    for c in corrections:
                        st.warning(c)
                else:
                    st.warning(str(corrections))

            if violations:
                st.markdown("##### Business Rule Fixes")
                for v in violations:
                    st.caption(f"⚠️ {v}")

            # Raw JSON expander
            with st.expander("Raw JSON"):
                st.json(raw)

# ── Run Tab ──────────────────────────────────────────────────────────────────
with tab_run:
    st.markdown("### Execute Pipeline")
    active_keys = _get_active_keys()
    provider = st.session_state.get("llm_provider", LLM_PROVIDER)
    if not active_keys:
        st.error("Add a Groq API key in the sidebar to run.")
        st.stop()

    _provider_labels = {"groq-8b-focused": "Groq-8b-Focused",
                        "groq-70b-focused": "Groq-70b-Focused"}
    provider_label = _provider_labels.get(provider, provider)
    st.success(f"{len(active_keys)} {provider_label} key(s) active")
    run_mode = st.radio("Mode", ["Single PDF", "Filter", "All rows"],
                        horizontal=True)

    run_file = run_brand = None
    run_n = None
    selected_pairs = None
    if run_mode == "Single PDF":
        rc1, rc2 = st.columns(2)
        with rc1:
            run_file_tag = st.selectbox("PDF", all_file_tags, key="run_f")
            run_file = _pdf_from_tag(run_file_tag)
        with rc2:
            run_brands = [b for f, b in submission_rows if f == run_file]
            run_brand = st.selectbox("Brand", run_brands, key="run_b")
    elif run_mode == "Filter":
        # Brand filter
        all_brands = sorted(set(b for _, b in submission_rows))
        selected_brands = st.multiselect(
            "Filter by brand", all_brands,
            placeholder="All brands (type to search)",
            key="run_brand_filter")

        # File+brand pair picker (filtered by brand selection)
        if selected_brands:
            available_pairs = [(f, b) for f, b in submission_rows
                               if b in selected_brands]
        else:
            available_pairs = list(submission_rows)

        def _pair_label(f, b):
            icon = "✅" if (f, b) in _cached_pairs else "◯"
            return f"{b} — {f}  [{icon}]"

        pair_labels = [_pair_label(f, b) for f, b in available_pairs]
        selected_labels = st.multiselect(
            "Select specific rows", pair_labels,
            placeholder="All matching rows (type to search)",
            key="run_pair_filter")

        if selected_labels:
            selected_pairs = [available_pairs[pair_labels.index(l)]
                              for l in selected_labels]
        elif selected_brands:
            selected_pairs = available_pairs

        # Status summary
        if selected_pairs:
            brands_in = sorted(set(b for _, b in selected_pairs))
            st.info(f"{len(selected_pairs)} row(s) selected — "
                    f"{', '.join(brands_in)}")
        else:
            st.caption(f"No filter applied — will run all "
                       f"{len(submission_rows)} rows")

    oc1, oc2 = st.columns(2)
    with oc1:
        skip_val = st.checkbox("Skip validation pass", help="Saves API calls")
    with oc2:
        use_cache = st.checkbox("Use cache", value=True)

    st.divider()

    if st.button("Run Pipeline", type="primary"):
        if run_mode == "Single PDF" and run_file:
            rows = [(f, b) for f, b in submission_rows if f == run_file]
            if run_brand:
                rows = [(f, b) for f, b in rows if b == run_brand]
        elif run_mode == "Filter" and selected_pairs:
            rows = selected_pairs
        else:
            rows = list(submission_rows)

        api_key = ",".join(active_keys)
        if provider == "groq-70b-focused":
            ext = OptimizedGroqExtractor(api_key=api_key, use_cache=use_cache)
        else:
            ext = StandardExtractor(api_key=api_key, use_cache=use_cache,
                                  provider="groq",
                                  pdf_dir=get_pdf_dir())

        is_single = len(rows) == 1

        if is_single:
            # ── Detailed single-row pipeline view ────────────────────
            fname, brand = rows[0]
            st.markdown(f"**{brand}** — `{fname}`")

            stages = [
                ("Read PDF", "Reading and parsing PDF pages..."),
                ("Filter Sections", "Locating relevant policy sections..."),
                ("Analyze Policy", f"Extracting 13 PA parameters via {provider_label}..."),
            ]
            if not skip_val:
                stages.append(("Validate", f"Cross-checking values via {provider_label}..."))
            stages.append(("Score", "Computing Access Score..."))

            stage_container = st.container()
            progress = st.progress(0)
            step_cols = stage_container.columns(len(stages))
            step_placeholders = []
            for i, (label, _) in enumerate(stages):
                with step_cols[i]:
                    ph = st.empty()
                    ph.markdown(f"<div style='text-align:center;color:#999;font-size:0.85em'>"
                                f"{label}</div>", unsafe_allow_html=True)
                    step_placeholders.append(ph)

            def _mark_stage(idx, status_icon="running"):
                icons = {"pending": "\u25cb", "running": "\u25f7", "done": "\u25cf", "error": "\u2716"}
                colors = {"pending": "#999", "running": "#1f77b4", "done": "#2ca02c", "error": "#d62728"}
                icon = icons.get(status_icon, "\u25cb")
                color = colors.get(status_icon, "#999")
                label = stages[idx][0]
                step_placeholders[idx].markdown(
                    f"<div style='text-align:center;color:{color};font-size:0.85em'>"
                    f"{icon} {label}</div>", unsafe_allow_html=True)

            detail = st.empty()

            # Stage 1: PDF extraction
            _mark_stage(0, "running")
            detail.caption(stages[0][1])
            filter_lvl = st.session_state.get("filter_level", "sentence")
            pkgs = process_all_pdfs(rows, filter_level=filter_lvl)
            pkg = pkgs[0]
            _mark_stage(0, "done")
            progress.progress(0.15)

            # Stage 2: Context filtering details
            _mark_stage(1, "running")
            detail.caption(f"{stages[1][1]}  |  {pkg.document_type}, "
                           f"{len(pkg.full_relevant_text):,} chars, "
                           f"filter: {filter_lvl}")
            import time as _time; _time.sleep(0.3)  # brief pause so user sees the info
            _mark_stage(1, "done")
            progress.progress(0.25)

            # Stage 3: LLM extraction
            _mark_stage(2, "running")
            detail.caption(stages[2][1])
            try:
                result = ext.extract_single(pkg)
            except AllKeysExhausted as exc:
                _mark_stage(2, "error")
                st.error(str(exc))
                st.stop()
            extractions = [result]
            _mark_stage(2, "done")
            progress.progress(0.6 if not skip_val else 0.85)

            # Stage 4: LLM validation (optional)
            if not skip_val:
                _mark_stage(3, "running")
                detail.caption(stages[3][1])
                try:
                    validated_result = ext.validate_single(result, pkg)
                except AllKeysExhausted as exc:
                    _mark_stage(3, "error")
                    st.error(str(exc))
                    st.stop()
                validated = [validated_result]
                _mark_stage(3, "done")
                progress.progress(0.85)
            else:
                validated = extractions
                # Remove stale validate cache so Results tab picks up
                # the fresh extract file (load_cached_results checks
                # validate before extract).
                provider = st.session_state.get("llm_provider", LLM_PROVIDER)
                stale_key = _cache_key(fname, brand, "validate", provider)
                stale_path = Path(CACHE_DIR) / stale_key
                if stale_path.exists():
                    stale_path.unlink()
                    print(f"  [{brand}] Removed stale validate cache")

            # Stage 5: Scoring
            score_idx = len(stages) - 1
            _mark_stage(score_idx, "running")
            detail.caption(stages[score_idx][1])
            final = []
            confidences = []
            for p, data in zip(pkgs, validated):
                row = normalize_output(data, p.filename, p.brand)
                row, _ = enforce_business_rules(row)
                score = compute_access_score(data, p.brand)
                row["Access Score"] = score
                confidences.append(compute_confidence(score, data))

                # Score sanity check
                sw = check_score_sanity(data, score, p.brand)
                for w in sw:
                    ext._log_manual_review(p.filename, p.brand,
                                           f"Score sanity: {w}")
                final.append(row)

            ext._flush_manual_review()
            _mark_stage(score_idx, "done")
            progress.progress(1.0)

            detail.empty()

            # Store results for display
            st.session_state["run_results"] = final
            st.session_state["run_confidences"] = confidences
            st.session_state["run_stats"] = {
                "calls": ext.api_calls_made,
                "cache_hits": ext.cache_hits,
                "input_tokens": ext.total_input_chars // 4,
                "output_tokens": ext.total_output_chars // 4,
                "provider": provider_label,
                "rows": len(final),
                "mode": "single",
            }
            st.rerun()

        else:
            # ── Batch mode (multiple rows) ───────────────────────────
            total = len(rows)
            st.info(f"Processing {total} row(s) via {provider_label}")

            # Layout: progress bar + current status + live log
            progress = st.progress(0)
            col_status, col_stats = st.columns([3, 1])
            status = col_status.empty()
            live_stats = col_stats.empty()
            log_container = st.container()
            log_area = log_container.empty()
            log_lines = []

            def _log(msg):
                log_lines.append(msg)
                # Show last 12 lines
                log_area.code("\n".join(log_lines[-12:]), language=None)

            def _update_stats():
                calls = ext.api_calls_made
                hits = ext.cache_hits
                esc = getattr(ext, 'escalations', 0)
                dead = len(getattr(ext, '_dead_keys', set()))
                live_stats.markdown(
                    f"**Calls:** {calls}  \n"
                    f"**Cache:** {hits}  \n"
                    f"**Dead keys:** {dead}")

            status.text("Reading PDFs & filtering relevant sections...")
            pkgs = process_all_pdfs(rows,
                                    filter_level=st.session_state.get("filter_level", "sentence"))
            _log(f"PDF reading done: {len(pkgs)} documents prepared")
            progress.progress(0.1)

            extractions = []
            for i, pkg in enumerate(pkgs):
                row_label = f"[{i+1}/{total}] {pkg.brand} — {pkg.filename}"
                status.text(f"Analyzing policy: {row_label}")

                # Capture pre-extraction state
                calls_before = ext.api_calls_made
                hits_before = ext.cache_hits

                try:
                    r = ext.extract_single(pkg)
                except AllKeysExhausted as exc:
                    st.error(str(exc))
                    if extractions:
                        st.warning(f"Partial results: {len(extractions)}/{total} rows "
                                   f"completed before keys ran out.")
                    st.stop()
                extractions.append(r)

                # Determine what happened
                calls_used = ext.api_calls_made - calls_before
                was_cached = ext.cache_hits > hits_before
                text_len = len(pkg.full_relevant_text)

                if was_cached:
                    _log(f"{row_label}: cached")
                else:
                    # Extract key results for display
                    brands = r.get("steps_through_brands", "?")
                    generic = r.get("steps_through_generic", "?")
                    photo = r.get("step_through_phototherapy", "?")
                    _log(f"{row_label}: {calls_used} calls, "
                         f"{text_len:,} chars | "
                         f"brands={brands} generic={generic} photo={photo}")

                _update_stats()
                progress.progress(0.1 + 0.4 * (i + 1) / total)

            if not skip_val:
                validated = []
                for i, (pkg, ex) in enumerate(zip(pkgs, extractions)):
                    row_label = f"[{i+1}/{total}] {pkg.brand}"
                    status.text(f"Validating: {row_label}")
                    try:
                        v = ext.validate_single(ex, pkg)
                    except AllKeysExhausted as exc:
                        st.error(str(exc))
                        st.warning("Validation stopped. Using unvalidated extractions "
                                   "for remaining rows.")
                        validated.extend(extractions[i:])
                        break
                    validated.append(v)

                    corrections = v.get("corrections")
                    if corrections:
                        _log(f"  {row_label}: {len(corrections)} correction(s)")

                    progress.progress(0.5 + 0.3 * (i + 1) / total)
            else:
                validated = extractions
                # Remove stale validate caches so Results tab picks up
                # fresh extract files.
                provider = st.session_state.get("llm_provider", LLM_PROVIDER)
                for pkg in pkgs:
                    stale_key = _cache_key(pkg.filename, pkg.brand,
                                           "validate", provider)
                    stale_path = Path(CACHE_DIR) / stale_key
                    if stale_path.exists():
                        stale_path.unlink()

            status.text("Scoring...")
            final = []
            confidences = []
            score_warning_count = 0
            for pkg, data in zip(pkgs, validated):
                row = normalize_output(data, pkg.filename, pkg.brand)
                row, _ = enforce_business_rules(row)
                score = compute_access_score(data, pkg.brand)
                row["Access Score"] = score
                confidences.append(compute_confidence(score, data))

                # Score sanity check
                sw = check_score_sanity(data, score, pkg.brand)
                for w in sw:
                    ext._log_manual_review(pkg.filename, pkg.brand,
                                           f"Score sanity: {w}")
                score_warning_count += len(sw)

                final.append(row)

            # Flush review flags (replaces file with only this run's entries)
            ext._flush_manual_review()

            done_msg = (f"Done: {total} rows, {ext.api_calls_made} API calls, "
                        f"{ext.cache_hits} cache hits")
            if score_warning_count:
                done_msg += f", {score_warning_count} score warning(s)"
            _log(done_msg)
            progress.progress(1.0)
            status.empty()

            # Store results for display
            st.session_state["run_results"] = final
            st.session_state["run_confidences"] = confidences
            st.session_state["run_stats"] = {
                "calls": ext.api_calls_made,
                "cache_hits": ext.cache_hits,
                "input_tokens": ext.total_input_chars // 4,
                "output_tokens": ext.total_output_chars // 4,
                "provider": provider_label,
                "rows": len(final),
                "mode": "batch",
            }
            st.rerun()

    # ── Display last run results (persists across reruns) ────────────
    if "run_results" in st.session_state and st.session_state["run_results"]:
        final = st.session_state["run_results"]
        stats = st.session_state.get("run_stats", {})
        df_run = pd.DataFrame(final)

        st.divider()
        st.markdown("### Last Run Results")

        # Summary metrics
        rc1, rc2, rc3, rc4 = st.columns(4)
        rc1.metric("Rows", stats.get("rows", len(final)))
        rc2.metric("API Calls", stats.get("calls", "?"))
        rc3.metric("Cache Hits", stats.get("cache_hits", 0))

        # Score distribution
        if "Access Score" in df_run.columns:
            scores = pd.to_numeric(df_run["Access Score"], errors="coerce")
            valid_scores = scores.dropna()
            if len(valid_scores) > 0:
                rc4.metric("Avg Score", f"{valid_scores.mean():.0f}")

                # Score breakdown
                st.markdown("##### Score Distribution")
                score_counts = valid_scores.value_counts().sort_index()
                sc_cols = st.columns(len(score_counts))
                for j, (score_val, count) in enumerate(score_counts.items()):
                    sc_cols[j].metric(f"Score {int(score_val)}", count)

        # Confidence check
        confidences = st.session_state.get("run_confidences", [])
        if confidences:
            valid_conf = [c for c in confidences if c["level"] != "N/A"]
            if valid_conf:
                with st.expander(
                    f"Extraction Confidence ({len(valid_conf)} rows scored)",
                    expanded=False
                ):
                    st.caption(
                        "Compares the LLM's own access score estimate against "
                        "the deterministic scorer. Large disagreements suggest "
                        "the extraction may have errors."
                    )
                    conf_rows = []
                    for i, c in enumerate(confidences):
                        if c["level"] == "N/A":
                            continue
                        row = final[i] if i < len(final) else {}
                        conf_rows.append({
                            "Brand": row.get("Brand", "?"),
                            "Filename": row.get("Filename", "?"),
                            "LLM Estimate": c["llm_estimate"],
                            "Deterministic": c["deterministic"],
                            "Diff": c["diff"],
                            "Confidence": c["level"],
                        })
                    df_conf = pd.DataFrame(conf_rows)

                    # Summary counts
                    level_counts = df_conf["Confidence"].value_counts()
                    cc1, cc2, cc3 = st.columns(3)
                    cc1.metric("High", level_counts.get("High", 0))
                    cc2.metric("Medium", level_counts.get("Medium", 0))
                    cc3.metric("Low", level_counts.get("Low", 0))

                    # Show only Medium/Low rows (High = no action needed)
                    flagged = df_conf[df_conf["Confidence"] != "High"]
                    if len(flagged) > 0:
                        st.markdown("**Rows to review:**")
                        safe_st_dataframe(flagged, width='stretch',
                                          hide_index=True)
                    else:
                        st.success("All rows have high confidence.")

        # Results table
        st.markdown("##### Extracted Parameters")
        safe_st_dataframe(df_run, width='stretch')

        # Per-row detail (expandable)
        if len(final) <= 10:
            with st.expander("Row Details", expanded=False):
                for row in final:
                    brand = row.get("Brand", "?")
                    fname = row.get("Filename", "?")
                    score = row.get("Access Score", "?")
                    st.markdown(f"**{brand}** — `{fname}` — Score: **{score}**")
                    detail_cols = {k: v for k, v in row.items()
                                   if k not in ("Filename", "Brand", "Access Score")}
                    st.json(detail_cols)

        # Save token usage stats for sidebar display
        st.session_state["last_run_stats"] = {
            "calls": stats.get("calls", 0),
            "cache_hits": stats.get("cache_hits", 0),
            "input_tokens": stats.get("input_tokens", 0),
            "output_tokens": stats.get("output_tokens", 0),
            "provider": stats.get("provider", "?"),
        }

        # Clear button
        if st.button("Clear Results"):
            for key in ("run_results", "run_stats", "run_confidences"):
                st.session_state.pop(key, None)
            st.rerun()

# ── Guide Tab ────────────────────────────────────────────────────────────────
with tab_guide:
    st.markdown("### Getting Started")
    st.markdown(
        "This dashboard lets you extract Prior Authorization policy parameters "
        "from payer PDFs, inspect every step of the extraction, and download "
        "structured results — all from your browser."
    )

    st.divider()

    st.markdown("#### 1. Add Your API Key")
    st.markdown(
        "Open the **sidebar** (left panel) and paste your Groq API key. "
        "You can get a free key at [console.groq.com](https://console.groq.com). "
        "Multiple keys can be added for faster processing — the pipeline "
        "rotates between them automatically to stay within rate limits."
    )

    st.markdown("#### 2. Run Extractions")
    st.markdown(
        "Go to the **Run** tab. You have three modes:\n\n"
        "- **Single PDF** — pick one file and brand from the dropdowns. "
        "Ideal for testing or re-running a specific row. You'll see a "
        "live stage-by-stage progress view.\n"
        "- **Filter** — search by brand name or select specific file+brand "
        "pairs. Cache status icons show which rows already have results.\n"
        "- **All rows** — process the entire dataset. Progress bar and live "
        "log keep you updated.\n\n"
        "Each extraction is cached automatically, so re-running a row is "
        "instant unless you uncheck **Use Cache**."
    )

    st.markdown("#### 3. Explore Results")
    st.markdown(
        "The **Results** tab shows a filterable table of all extracted parameters. "
        "Select any row to inspect its values and see a detailed score breakdown "
        "showing how each parameter contributes to the Access Score.\n\n"
        "Download the full dataset as **CSV** or **Excel** using the buttons "
        "at the bottom of the table."
    )

    st.markdown("#### 4. Inspect PDFs in Detail")
    st.markdown(
        "The **PDF Explorer** tab lets you deep-dive into any PDF:\n\n"
        "- **Page relevance scores** — see which pages the pipeline identified "
        "as containing PA criteria and why.\n"
        "- **Context tiers** — view the exact text sent to the LLM at each "
        "filtering level (page → paragraph → sentence).\n"
        "- **Prompt preview** — see the full prompt that would be sent, "
        "including few-shot examples.\n"
        "- **Raw extraction JSON** — the LLM's raw output for debugging.\n\n"
        "This is useful for understanding why a particular value was extracted "
        "or for verifying the pipeline's interpretation of a complex policy."
    )

    st.markdown("#### 5. Review the Overview")
    st.markdown(
        "The **Overview** tab provides a high-level summary once results are "
        "available:\n\n"
        "- **Access Score Distribution** — how policies spread across the "
        "five score tiers.\n"
        "- **Brand Comparison** — which brands face more or fewer PA barriers "
        "on average.\n"
        "- **Step Therapy Burden** — the number of drugs patients must try "
        "and fail before getting their prescribed biologic approved."
    )

    st.divider()

    st.markdown("#### Tips")
    st.markdown(
        "- **Multiple API keys** speed things up significantly. The free Groq "
        "tier has per-key rate limits, so 3-4 keys let the pipeline rotate "
        "and avoid cooldown pauses.\n"
        "- **Context filter level** (sidebar) controls how aggressively PDF "
        "text is trimmed before sending to the LLM. *Sentence* (default) is "
        "most aggressive and works well for most PDFs. Try *paragraph* if "
        "a specific extraction looks incomplete.\n"
        "- **Skip validation** (Run tab checkbox) halves the API calls per row. "
        "Useful for quick exploratory runs.\n"
        "- The pipeline handles diverse PDF formats — single-drug policies, "
        "multi-drug formularies, decision trees, and flat catalogs — with "
        "format-specific context filtering."
    )

    st.divider()
    st.caption(
        "Pipeline: PDF ingestion → context filtering → LLM extraction → "
        "validation → Access Score computation."
    )


